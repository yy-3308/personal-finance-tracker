"""Parse American Express activity XLSX exports."""

import re
from datetime import datetime

import openpyxl

from categorizer import categorize

# Map AMEX categories to our categories
_AMEX_CAT_MAP = {
    "Merchandise & Supplies-Groceries": "Groceries",
    "Restaurant-Restaurant": "Dining",
    "Restaurant-Bar & Café": "Dining",
    "Transportation-Fuel": "Gas",
    "Travel-Travel Agencies": "Travel",
    "Travel-Airline": "Travel",
    "Travel-Lodging": "Travel",
    "Travel-Other": "Travel",
    "Entertainment-General Attractions": "Entertainment",
    "Business Services-Advertising Services": "Subscription",
    "Fees & Adjustments-Fees & Adjustments": "Fee",
}


def is_amex_xlsx(filepath):
    """Check if an XLSX file is an AMEX activity export."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb["Transaction Details"]
        title = str(ws.cell(1, 2).value or "") + str(ws.cell(1, 1).value or "")
        wb.close()
        return "Blue Cash" in title or "Gold Card" in title or "Platinum" in title
    except Exception:
        return False


def parse_amex_xlsx(filepath):
    """Parse an AMEX activity XLSX. Returns card info and transactions."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Transaction Details"]

    # Extract card name and account number from header rows
    title = str(ws.cell(1, 2).value or "")
    card_name = "Amex Card"
    card_match = re.search(r"(Blue Cash Everyday|Gold Card|Platinum|Delta SkyMiles)", title)
    if card_match:
        card_name = f"Amex {card_match.group(1)}"

    acct_str = str(ws.cell(5, 1).value or "")
    last5 = re.search(r"(\d{5})$", acct_str)
    last_digits = last5.group(1) if last5 else ""

    # Parse transactions starting after header row (row 7)
    transactions = []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        date_val = row[0]
        desc_val = row[1]
        amount_val = row[2]
        amex_category = str(row[10] or "") if len(row) > 10 else ""

        if date_val is None or amount_val is None:
            continue

        # Parse date
        if isinstance(date_val, datetime):
            txn_date = date_val.strftime("%Y-%m-%d")
        else:
            try:
                txn_date = datetime.strptime(str(date_val), "%m/%d/%Y").strftime("%Y-%m-%d")
            except ValueError:
                continue

        # Parse amount
        try:
            amount = float(amount_val)
        except (ValueError, TypeError):
            continue

        # Clean description
        description = str(desc_val or "").strip()
        # Remove "AplPay " prefix and clean up spacing
        description = re.sub(r"^AplPay\s+", "", description)
        # Collapse extra whitespace
        description = re.sub(r"\s{2,}", " ", description).strip()

        # Skip autopay (transfer between accounts)
        if "AUTOPAY" in description.upper():
            continue

        # Map category
        category = _AMEX_CAT_MAP.get(amex_category)
        if not category:
            category = categorize(description)

        # Flip sign: AMEX positive = charge (expense), negative = credit/refund
        amount = -amount

        transactions.append({
            "date": txn_date,
            "description": description,
            "amount": amount,
            "category": category,
        })

    wb.close()

    # Determine month range from transactions
    months = set()
    for t in transactions:
        months.add(t["date"][:7])

    return {
        "type": "amex_credit_card",
        "card_name": card_name,
        "last_digits": last_digits,
        "transactions": transactions,
        "months": sorted(months),
    }
